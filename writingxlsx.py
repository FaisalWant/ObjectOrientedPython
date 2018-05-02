from openpyxl import Workbook
import datetime
from datetime import timedelta, date

  
def daterange(start_date, end_date):
    for n in range(int ((end_date - start_date).days)):
        yield start_date + timedelta(n)


# initializing the header of the Excel file


start_date = date(2013, 1, 1)
end_date = date(2015, 6, 2)

wb = Workbook()
wb.create_sheet('Sample')
ws = wb.active
ws['A1'] = "id"
ws['B1'] = "Date"
ws['C1'] = "year"
# ws['C1'] = "password"
# ws['D1'] = "Date"

for single_date in daterange(start_date, end_date):
    print (single_date.strftime("%Y-%m-%d"))
    takeme=str(single_date.strftime("%Y-%m-%d"))
    mylist=takeme.split("-")
   
    ws.append(["date",takeme, "QBT"+str(mylist[0])])

# Can use Python datetime objects

# Change sheet tab color
ws.sheet_properties.tabColor = "660000"

wb.save("users.xlsx")  # Write to disk