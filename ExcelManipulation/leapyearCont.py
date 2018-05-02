#leapyearCont.py

from datetime import datetime
from datetime import timedelta, date
import datetime
import calendar
from openpyxl import Workbook

calen={"1":31,"2":28,"3":31,"4":30, "5":31, "6":30,"7":31,
"8":31, "9":30, "10":31,"11":30,"12":31}

#777777777777777777777777777777777777777777777777777777777777777777777
def month_year_iter( start_month, start_year, end_month, end_year ):
	ym_start= 12*start_year + start_month -1
	ym_end= 12*end_year + end_month 
	for ym in range( ym_start, ym_end ):
		y, m = divmod( ym, 12 )
		yield y, m+1


start=datetime.date(2011, 5, 5)
end=datetime.date(2016, 5, 6)

sta_mth=start.month
sta_yea=start.year
sta_day=start.day

end_mth=end.month
end_yea=end.year
end_day=end.day
count=0;
# creating the Excel workbook
wb = Workbook()
wb.create_sheet('Sample')
ws = wb.active
ws['A1'] = "Startdate"
ws['B1'] = "EndDate"
ws['C1'] = "year"
# populating the Excel file
 
for i in month_year_iter(sta_mth,sta_yea,end_mth,end_yea):
	data = [str(x) for x in i]
	str1='-'.join(data)
	yea=str(str1.split("-")[0])
	mon=str(str1.split("-")[1])
	# checking tail 
	if (mon==str(end_mth))and(yea == str(end_yea)):
		print("*" * 3)
		# print
		print(str1+"-"+str(end_day))
		takeFirst= str1+"-"+str(end_day)   # printing Start
		str_end=str1+"-"+str(end_day)
		if(calendar.isleap(int(str_firs.split("-")[0]))) and (str_firs.split("-")[1]=="2"):        
			print(str_end.split("-")[0]+"-"+str_firs.split("-")[1]+"-"+"29")
			end_me=str_end.split("-")[0]+"-"+str_firs.split("-")[1]+"-"+"29"
			ws.append([takeFirst , end_me, "QBTB"+end_me.split("-")[0]+""+end_me.split("-")[1]])
		else:
			print(str_end.split("-")[0]+"-"+str_end.split("-")[1]+"-"+str(calen.get(str_end.split("-")[1])))
			end_me1=str_end.split("-")[0]+"-"+str_end.split("-")[1]+"-"+str(calen.get(str_end.split("-")[1]))
			ws.append([takeFirst , end_me1,"QBTB"+end_me1.split("-")[0]+""+end_me1.split("-")[1] ])
	else:
		if count==0:
			print(str1+"-"+str(sta_day))  # printing  start date
			str_firs=str1+"-"+str(sta_day)
			if (calendar.isleap(int(str_firs.split("-")[0]))) and (str_firs.split("-")[1]=="2"):
				print(str_firs.split("-")[0]+"-"+str_firs.split("-")[1]+"-"+"29")
				my_lif=str_firs.split("-")[0]+"-"+str_firs.split("-")[1]+"-"+"29"
				ws.append([str_firs , my_lif, "QBTB"+my_lif.split("-")[0]+""+my_lif.split("-")[1]])
			else:
				print(str_firs.split("-")[0]+"-"+str_firs.split("-")[1]+"-"+str(calen.get(str_firs.split("-")[1])))	
				whois=str_firs.split("-")[0]+"-"+str_firs.split("-")[1]+"-"+str(calen.get(str_firs.split("-")[1]))	
				ws.append([str_firs , whois, "QBTB"+whois.split("-")[0]+""+whois.split("-")[1]])		
			print("*" * 3)
			count=1	
		

		else:
			str2=str1+"-"+"1"
			print(str2)                  # printing start date
			sec_dat=str2.split("-")[1]
			print("Display the Other End")
			if (calendar.isleap(int(str2.split("-")[0]))) and (str2.split("-")[1]=="2") :

				print(str2.split("-")[0]+"-"+str2.split("-")[1]+"-"+"29")
				ilove=str2.split("-")[0]+"-"+str2.split("-")[1]+"-"+"29"
				ws.append([str2 , ilove, "QBTB"+ilove.split("-")[0]+""+ilove.split("-")[1]])
			else:
				print(str2.split("-")[0]+"-"+str2.split("-")[1]+"-"+str(calen.get(sec_dat)))
				youlove=str2.split("-")[0]+"-"+str2.split("-")[1]+"-"+str(calen.get(sec_dat))
				ws.append([str2 , youlove,"QBTB"+youlove.split("-")[0]+""+youlove.split("-")[1] ])
					  
									
ws.sheet_properties.tabColor = "660000"

wb.save("khan.xlsx")  # Write to disk

		
	
